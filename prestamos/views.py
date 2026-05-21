# prestamos/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
from datetime import timedelta, date
from catalogo.models import Libro
from .models import Reserva, Prestamo, Sancion
from .forms import FormularioReserva, FormularioPrestamo, FiltroPrestamoForm

# Imports para exportación
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


# ─────────────── RESERVAS ───────────────

@login_required
def lista_reservas(request):
    """Lista de reservas. El bibliotecario ve todas; el lector solo las suyas."""
    if request.user.es_bibliotecario:
        reservas = Reserva.objects.select_related('usuario', 'libro').all()
    else:
        reservas = Reserva.objects.filter(usuario=request.user)

    return render(request, 'prestamos/lista_reservas.html', {'reservas': reservas})


@login_required
def crear_reserva(request, libro_pk):
    """Crear una reserva para un libro."""
    libro = get_object_or_404(Libro, pk=libro_pk)

    # Verificar que el usuario no esté sancionado
    if request.user.sancionado:
        messages.error(request, 'No puedes reservar libros mientras estés sancionado.')
        return redirect('catalogo:libro_detalle', pk=libro_pk)

    # Verificar disponibilidad
    if not libro.esta_disponible:
        messages.error(request, 'Este libro no está disponible para reserva.')
        return redirect('catalogo:libro_detalle', pk=libro_pk)

    # Verificar que no tenga una reserva activa del mismo libro
    reserva_existente = Reserva.objects.filter(
        usuario=request.user,
        libro=libro,
        estado__in=['pendiente', 'confirmada']
    ).exists()
    if reserva_existente:
        messages.warning(request, 'Ya tienes una reserva activa para este libro.')
        return redirect('catalogo:libro_detalle', pk=libro_pk)

    if request.method == 'POST':
        form = FormularioReserva(request.POST)
        if form.is_valid():
            reserva = form.save(commit=False)
            reserva.usuario = request.user
            reserva.libro = libro
            reserva.fecha_expiracion = timezone.now().date() + timedelta(days=3)
            reserva.save()

            # Actualizar estado del libro si es el único ejemplar
            if libro.ejemplares_disponibles == 1:
                libro.estado = Libro.ESTADO_RESERVADO
            libro.ejemplares_disponibles -= 1
            libro.save()

            messages.success(
                request,
                f'Reserva del libro "{libro.titulo}" creada. '
                f'Tienes hasta el {reserva.fecha_expiracion} para recogerlo.'
            )
            return redirect('prestamos:reservas')
    else:
        form = FormularioReserva()

    return render(request, 'prestamos/crear_reserva.html', {'form': form, 'libro': libro})


@login_required
def cancelar_reserva(request, pk):
    """Cancelar una reserva."""
    reserva = get_object_or_404(Reserva, pk=pk)

    # Solo el propio usuario o el bibliotecario pueden cancelar
    if request.user != reserva.usuario and not request.user.es_bibliotecario:
        messages.error(request, 'No tienes permiso para cancelar esta reserva.')
        return redirect('prestamos:reservas')

    if request.method == 'POST':
        reserva.estado = Reserva.ESTADO_CANCELADA
        reserva.save()

        # Devolver el ejemplar
        libro = reserva.libro
        libro.ejemplares_disponibles += 1
        if libro.ejemplares_disponibles > 0:
            libro.estado = Libro.ESTADO_DISPONIBLE
        libro.save()

        messages.success(request, 'Reserva cancelada correctamente.')
        return redirect('prestamos:reservas')

    return render(request, 'prestamos/cancelar_reserva.html', {'reserva': reserva})


# ─────────────── PRÉSTAMOS ───────────────

@login_required
def lista_prestamos(request):
    """Lista de préstamos con filtros."""
    form_filtro = FiltroPrestamoForm(request.GET)

    if request.user.es_bibliotecario:
        prestamos = Prestamo.objects.select_related('usuario', 'libro').all()
    else:
        prestamos = Prestamo.objects.filter(usuario=request.user)

    if form_filtro.is_valid():
        estado = form_filtro.cleaned_data.get('estado')
        fecha_desde = form_filtro.cleaned_data.get('fecha_desde')
        fecha_hasta = form_filtro.cleaned_data.get('fecha_hasta')
        q = form_filtro.cleaned_data.get('q')

        if estado:
            prestamos = prestamos.filter(estado=estado)
        if fecha_desde:
            prestamos = prestamos.filter(fecha_prestamo__date__gte=fecha_desde)
        if fecha_hasta:
            prestamos = prestamos.filter(fecha_prestamo__date__lte=fecha_hasta)
        if q:
            prestamos = prestamos.filter(
                Q(usuario__username__icontains=q) |
                Q(usuario__first_name__icontains=q) |
                Q(libro__titulo__icontains=q)
            )

    # Actualizar estados retrasados automáticamente
    hoy = timezone.now().date()
    prestamos_activos_retrasados = prestamos.filter(
        estado=Prestamo.ESTADO_ACTIVO,
        fecha_vencimiento__lt=hoy
    )
    prestamos_activos_retrasados.update(estado=Prestamo.ESTADO_RETRASADO)

    return render(request, 'prestamos/lista_prestamos.html', {
        'prestamos': prestamos,
        'form_filtro': form_filtro,
    })


@login_required
def crear_prestamo(request, reserva_pk=None):
    """Crear un préstamo (solo bibliotecarios). Puede venir de una reserva."""
    if not request.user.es_bibliotecario:
        messages.error(request, 'Solo el bibliotecario puede registrar préstamos.')
        return redirect('prestamos:lista')

    reserva = None
    libro = None

    if reserva_pk:
        reserva = get_object_or_404(Reserva, pk=reserva_pk)
        libro = reserva.libro
        usuario = reserva.usuario
    else:
        libro_pk = request.GET.get('libro')
        if libro_pk:
            libro = get_object_or_404(Libro, pk=libro_pk)

    if request.method == 'POST':
        form = FormularioPrestamo(request.POST)
        if form.is_valid():
            prestamo = form.save(commit=False)
            prestamo.libro = libro
            prestamo.usuario = usuario if reserva else request.user  # En producción: seleccionar usuario
            prestamo.reserva = reserva
            prestamo.save()

            # Actualizar reserva
            if reserva:
                reserva.estado = Reserva.ESTADO_CONFIRMADA
                reserva.save()

            # Actualizar libro
            libro.estado = Libro.ESTADO_PRESTADO
            libro.save()

            messages.success(
                request,
                f'Préstamo registrado. Vence el {prestamo.fecha_vencimiento}.'
            )
            return redirect('prestamos:lista')
    else:
        form = FormularioPrestamo()

    return render(request, 'prestamos/crear_prestamo.html', {
        'form': form,
        'reserva': reserva,
        'libro': libro,
    })


@login_required
def devolver_libro(request, pk):
    """Registrar la devolución de un libro."""
    if not request.user.es_bibliotecario:
        messages.error(request, 'Solo el bibliotecario puede registrar devoluciones.')
        return redirect('prestamos:lista')

    prestamo = get_object_or_404(Prestamo, pk=pk)

    if request.method == 'POST':
        prestamo.fecha_devolucion = timezone.now()

        hoy = timezone.now().date()
        if hoy > prestamo.fecha_vencimiento:
            dias_retraso = (hoy - prestamo.fecha_vencimiento).days
            prestamo.dias_retraso = dias_retraso
            prestamo.estado = Prestamo.ESTADO_RETRASADO

            # Crear sanción (1 día de sanción por cada día de retraso)
            dias_sancion = dias_retraso * 2  # 2 días de sanción por día de retraso
            fecha_fin_sancion = hoy + timedelta(days=dias_sancion)

            Sancion.objects.create(
                usuario=prestamo.usuario,
                prestamo=prestamo,
                fecha_fin=fecha_fin_sancion,
                dias_sancion=dias_sancion,
                motivo=f'Retraso de {dias_retraso} día(s) en la devolución de "{prestamo.libro.titulo}"'
            )

            # Sancionar al usuario
            prestamo.usuario.sancionado = True
            prestamo.usuario.fecha_fin_sancion = fecha_fin_sancion
            prestamo.usuario.save()

            messages.warning(
                request,
                f'Libro devuelto con {dias_retraso} días de retraso. '
                f'Se aplicó sanción hasta el {fecha_fin_sancion}.'
            )
        else:
            prestamo.estado = Prestamo.ESTADO_DEVUELTO
            messages.success(request, 'Libro devuelto exitosamente a tiempo.')

        prestamo.save()

        # Actualizar libro
        libro = prestamo.libro
        libro.ejemplares_disponibles += 1
        if libro.ejemplares_disponibles > 0:
            libro.estado = Libro.ESTADO_DISPONIBLE
        libro.save()

        return redirect('prestamos:lista')

    return render(request, 'prestamos/devolver_libro.html', {'prestamo': prestamo})


@login_required
def detalle_prestamo(request, pk):
    prestamo = get_object_or_404(Prestamo, pk=pk)
    if request.user != prestamo.usuario and not request.user.es_bibliotecario:
        messages.error(request, 'No tienes permiso para ver este préstamo.')
        return redirect('prestamos:lista')
    return render(request, 'prestamos/detalle_prestamo.html', {'prestamo': prestamo})


@login_required
def lista_sanciones(request):
    """Lista de sanciones activas."""
    if request.user.es_bibliotecario:
        sanciones = Sancion.objects.select_related('usuario', 'prestamo').all()
    else:
        sanciones = Sancion.objects.filter(usuario=request.user)
    return render(request, 'prestamos/lista_sanciones.html', {'sanciones': sanciones})


# ─────────────── REPORTES ───────────────

def _aplicar_filtros_reporte(request):
    """Helper: aplica los filtros de FiltroPrestamoForm y devuelve (queryset, form)."""
    form = FiltroPrestamoForm(request.GET)
    prestamos = Prestamo.objects.select_related('usuario', 'libro').all()

    if form.is_valid():
        estado = form.cleaned_data.get('estado')
        fecha_desde = form.cleaned_data.get('fecha_desde')
        fecha_hasta = form.cleaned_data.get('fecha_hasta')
        q = form.cleaned_data.get('q')

        if estado:
            prestamos = prestamos.filter(estado=estado)
        if fecha_desde:
            prestamos = prestamos.filter(fecha_prestamo__date__gte=fecha_desde)
        if fecha_hasta:
            prestamos = prestamos.filter(fecha_prestamo__date__lte=fecha_hasta)
        if q:
            prestamos = prestamos.filter(
                Q(usuario__username__icontains=q) |
                Q(usuario__first_name__icontains=q) |
                Q(usuario__last_name__icontains=q) |
                Q(libro__titulo__icontains=q)
            )

    return prestamos.order_by('-fecha_prestamo'), form


@login_required
def reportes(request):
    """Vista principal de reportes — solo bibliotecarios."""
    if not request.user.es_bibliotecario:
        messages.error(request, 'Solo los bibliotecarios pueden acceder a los reportes.')
        return redirect('dashboard:index')

    prestamos, form_filtro = _aplicar_filtros_reporte(request)

    return render(request, 'prestamos/reportes.html', {
        'prestamos': prestamos,
        'form_filtro': form_filtro,
        'total': prestamos.count(),
    })


@login_required
def exportar_pdf(request):
    """Exporta los préstamos filtrados como PDF — solo bibliotecarios."""
    if not request.user.es_bibliotecario:
        return HttpResponse('Acceso denegado.', status=403)

    prestamos, _ = _aplicar_filtros_reporte(request)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=1.5 * cm,
    )

    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        'titulo',
        parent=estilos['Heading1'],
        alignment=TA_CENTER,
        fontSize=16,
        textColor=colors.HexColor('#5c3d2e'),
        spaceAfter=6,
    )
    estilo_sub = ParagraphStyle(
        'subtitulo',
        parent=estilos['Normal'],
        alignment=TA_CENTER,
        fontSize=9,
        textColor=colors.HexColor('#888888'),
        spaceAfter=14,
    )

    elementos = []
    elementos.append(Paragraph('Biblioteca Digital — Reporte de Préstamos', estilo_titulo))
    elementos.append(Paragraph(
        f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")} | Total: {prestamos.count()} préstamo(s)',
        estilo_sub
    ))

    # Encabezados
    cabecera = ['#', 'Libro', 'Usuario', 'Fecha préstamo', 'Vencimiento', 'Devolución', 'Estado', 'Retraso (días)']
    filas = [cabecera]

    ESTADOS = {
        'activo': 'Activo',
        'devuelto': 'Devuelto',
        'retrasado': 'Retrasado',
        'renovado': 'Renovado',
    }

    for i, p in enumerate(prestamos, start=1):
        devolucion = p.fecha_devolucion.strftime('%d/%m/%Y') if p.fecha_devolucion else '—'
        filas.append([
            str(i),
            p.libro.titulo[:40] + ('…' if len(p.libro.titulo) > 40 else ''),
            p.usuario.get_full_name() or p.usuario.username,
            p.fecha_prestamo.strftime('%d/%m/%Y'),
            p.fecha_vencimiento.strftime('%d/%m/%Y'),
            devolucion,
            ESTADOS.get(p.estado, p.estado),
            str(p.dias_retraso) if p.dias_retraso else '—',
        ])

    COLOR_HEADER = colors.HexColor('#5c3d2e')
    COLOR_FILA_PAR = colors.HexColor('#faf6f0')

    tabla = Table(filas, repeatRows=1, colWidths=[1*cm, 7*cm, 4.5*cm, 3.2*cm, 3.2*cm, 3.2*cm, 2.8*cm, 3*cm])
    estilo_tabla = TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), COLOR_HEADER),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        # Filas de datos
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (7, 1), (7, -1), 'CENTER'),
        # Filas alternas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_FILA_PAR]),
        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('LINEBELOW', (0, 0), (-1, 0), 1.2, COLOR_HEADER),
    ])
    tabla.setStyle(estilo_tabla)
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_prestamos.pdf"'
    return response


@login_required
def exportar_excel(request):
    """Exporta los préstamos filtrados como Excel — solo bibliotecarios."""
    if not request.user.es_bibliotecario:
        return HttpResponse('Acceso denegado.', status=403)

    prestamos, _ = _aplicar_filtros_reporte(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Préstamos'

    # Estilos
    color_cafe = 'FF5C3D2E'
    color_dorado = 'FFC9A84C'
    color_crema = 'FFFFF8F0'

    fuente_header = Font(bold=True, color='FFFFFFFF', size=10)
    fuente_titulo = Font(bold=True, color=color_cafe, size=13)
    relleno_header = PatternFill(fill_type='solid', fgColor=color_cafe)
    relleno_par = PatternFill(fill_type='solid', fgColor=color_crema)
    borde_celda = Border(
        left=Side(style='thin', color='FFCCCCCC'),
        right=Side(style='thin', color='FFCCCCCC'),
        top=Side(style='thin', color='FFCCCCCC'),
        bottom=Side(style='thin', color='FFCCCCCC'),
    )
    centrado = Alignment(horizontal='center', vertical='center')
    izquierda = Alignment(horizontal='left', vertical='center')

    # Título del reporte
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Biblioteca Digital — Reporte de Préstamos'
    ws['A1'].font = fuente_titulo
    ws['A1'].alignment = centrado
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    ws['A2'] = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")} | Total: {prestamos.count()} préstamo(s)'
    ws['A2'].font = Font(italic=True, color='FF888888', size=9)
    ws['A2'].alignment = centrado
    ws.row_dimensions[2].height = 18

    # Encabezados (fila 3)
    cabeceras = ['#', 'Libro', 'Usuario', 'Fecha préstamo', 'Vencimiento', 'Devolución', 'Estado', 'Retraso (días)']
    for col, texto in enumerate(cabeceras, start=1):
        celda = ws.cell(row=3, column=col, value=texto)
        celda.font = fuente_header
        celda.fill = relleno_header
        celda.alignment = centrado
        celda.border = borde_celda
    ws.row_dimensions[3].height = 22

    # Datos
    ESTADOS = {
        'activo': 'Activo',
        'devuelto': 'Devuelto',
        'retrasado': 'Retrasado',
        'renovado': 'Renovado',
    }

    for i, p in enumerate(prestamos, start=1):
        fila = i + 3
        devolucion = p.fecha_devolucion.strftime('%d/%m/%Y') if p.fecha_devolucion else '—'
        valores = [
            i,
            p.libro.titulo,
            p.usuario.get_full_name() or p.usuario.username,
            p.fecha_prestamo.strftime('%d/%m/%Y'),
            p.fecha_vencimiento.strftime('%d/%m/%Y'),
            devolucion,
            ESTADOS.get(p.estado, p.estado),
            p.dias_retraso if p.dias_retraso else '—',
        ]
        relleno_fila = relleno_par if i % 2 == 0 else None
        alineaciones = [centrado, izquierda, izquierda, centrado, centrado, centrado, centrado, centrado]

        for col, (valor, ali) in enumerate(zip(valores, alineaciones), start=1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.alignment = ali
            celda.border = borde_celda
            if relleno_fila:
                celda.fill = relleno_fila

        ws.row_dimensions[fila].height = 18

        # Color de estado
        celda_estado = ws.cell(row=fila, column=7)
        if p.estado == 'retrasado':
            celda_estado.font = Font(bold=True, color='FFC0392B')
        elif p.estado == 'activo':
            celda_estado.font = Font(bold=True, color='FF27AE60')
        elif p.estado == 'devuelto':
            celda_estado.font = Font(color='FF555555')

    # Anchos de columna
    anchos = [5, 42, 28, 16, 16, 16, 14, 16]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # Fila de totales
    fila_total = prestamos.count() + 4
    ws.merge_cells(f'A{fila_total}:G{fila_total}')
    ws[f'A{fila_total}'] = f'TOTAL DE PRÉSTAMOS: {prestamos.count()}'
    ws[f'A{fila_total}'].font = Font(bold=True, color=color_cafe, size=10)
    ws[f'A{fila_total}'].fill = PatternFill(fill_type='solid', fgColor='FFFEF3DC')
    ws[f'A{fila_total}'].alignment = centrado
    ws[f'A{fila_total}'].border = borde_celda
    ws.row_dimensions[fila_total].height = 20

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_prestamos.xlsx"'
    return response